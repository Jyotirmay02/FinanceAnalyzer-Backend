"""
Excel output writer for generating analysis reports with advanced formatting
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any
import logging
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Handles writing analysis results to Excel files"""
    
    def __init__(self, output_path: str):
        """
        Initialize Excel writer
        
        Args:
            output_path: Path for output Excel file
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
    
    def write_analysis_report(self, categorized_df: pd.DataFrame, 
                            overall_summary: Dict[str, Any],
                            category_summary: pd.DataFrame) -> None:
        """
        Write complete analysis report to Excel with multiple sheets
        
        Args:
            categorized_df: DataFrame with categorized transactions
            overall_summary: Overall financial summary dictionary
            category_summary: Category-wise summary DataFrame
        """
        try:
            # Remove existing file if it exists (overwrite)
            if self.output_path.exists():
                self.output_path.unlink()
                
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Write categorized transactions
                categorized_df.to_excel(
                    writer, 
                    sheet_name='Categorized Transactions', 
                    index=False
                )
                
                # Write overall summary
                summary_df = pd.DataFrame(
                    list(overall_summary.items()), 
                    columns=['Metric', 'Value']
                )
                summary_df.to_excel(
                    writer, 
                    sheet_name='Overall Summary', 
                    index=False
                )
                
                # Write category summary
                category_summary.to_excel(
                    writer, 
                    sheet_name='Category Summary', 
                    index=False
                )
                
                # TODO: Add Top 10 Transactions sheet (moved to enhancements)
                # top_transactions = self._get_top_transactions(categorized_df, 10)
                # top_transactions.to_excel(writer, sheet_name='Top 10 Transactions', index=False)
                
                # Add UPI Analysis if UPI subcategories exist
                upi_analysis = self._get_upi_construction_analysis(categorized_df)
                if len(upi_analysis) > 0:
                    upi_analysis.to_excel(
                        writer,
                        sheet_name='UPI Analysis',
                        index=False
                    )
                    # Format UPI Analysis sheet
                    self._format_upi_analysis_sheet(writer.book['UPI Analysis'], upi_analysis)
                
                # Add UPI Summary sheet with hierarchical grouping
                self._create_upi_hierarchical_sheet(categorized_df, writer)
                
                # TODO: Add Filter Controls sheet (temporarily disabled due to merged cell issue)
                # self._create_filter_controls_sheet(categorized_df, writer)
                
                # Apply formatting to all sheets
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Skip auto-fit for UPI Summary and Filter Controls (have custom formatting)
                    if sheet_name not in ['UPI Summary', 'Filter Controls']:
                        # Auto-fit columns with better width calculation
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = None
                            
                            for cell in column:
                                # Skip merged cells
                                if hasattr(cell, 'column_letter'):
                                    if column_letter is None:
                                        column_letter = cell.column_letter
                                    try:
                                        cell_length = len(str(cell.value))
                                        if cell_length > max_length:
                                            max_length = cell_length
                                    except:
                                        pass
                            
                            # Set column width if we have a valid column letter
                            if column_letter:
                                # Set column width with padding, minimum 12 for currency columns
                                adjusted_width = max(max_length + 3, 12)
                                adjusted_width = min(adjusted_width, 50)  # Cap at 50
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Apply header formatting
                    self._format_headers(worksheet)
                    
                    # Apply currency formatting and headers to relevant columns
                    if sheet_name in ['Categorized Transactions']:
                        # Find Debit, Credit, Balance columns dynamically
                        header_row = [cell.value for cell in worksheet[1]]
                        currency_cols = []
                        for idx, header in enumerate(header_row):
                            if header in ['Debit', 'Credit', 'Balance', 'Amount']:
                                currency_cols.append(worksheet.cell(row=1, column=idx+1).column_letter)
                        self._format_currency_columns(worksheet, currency_cols)
                        self._add_currency_headers(worksheet, currency_cols)
                    elif sheet_name == 'Category Summary':
                        # Find Total Debit, Total Credit columns
                        header_row = [cell.value for cell in worksheet[1]]
                        currency_cols = []
                        for idx, header in enumerate(header_row):
                            if 'Debit' in str(header) or 'Credit' in str(header):
                                currency_cols.append(worksheet.cell(row=1, column=idx+1).column_letter)
                        self._format_currency_columns(worksheet, currency_cols)
                        self._add_currency_headers(worksheet, currency_cols)
                    
                    # Add filters and freeze panes
                    if worksheet.max_row > 1:
                        max_col_letter = worksheet.cell(row=1, column=worksheet.max_column).column_letter
                        self._add_filters(worksheet, worksheet.max_row, max_col_letter)
                        self._freeze_panes(worksheet)
                    
                    # Add better visualization to Category Summary sheet (horizontal bar chart)
                    if sheet_name == 'Category Summary' and len(category_summary) > 0:
                        # Position chart 4 columns right from last used column (F -> J)
                        self._add_horizontal_bar_chart(worksheet, category_summary, "J5")
                
                logger.info(f"Analysis report written to {self.output_path}")
                
        except Exception as e:
            logger.error(f"Error writing Excel report: {e}")
            raise
                
    def _format_headers(self, worksheet, start_row=1):
        """Apply header formatting - bold, colored background"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[start_row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    def _format_currency_columns(self, worksheet, currency_columns):
        """Format currency columns with number formatting (no ₹ symbol)"""
        for col in currency_columns:
            for cell in worksheet[col]:
                if cell.row > 1 and cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    def _add_currency_headers(self, worksheet, currency_columns):
        """Add (₹) to currency column headers"""
        for col in currency_columns:
            header_cell = worksheet[f"{col}1"]
            if header_cell.value and not "(₹)" in str(header_cell.value):
                header_cell.value = f"{header_cell.value} (₹)"
    
    def _add_filters(self, worksheet, end_row, end_col):
        """Add Excel filters to data"""
        worksheet.auto_filter.ref = f"A1:{end_col}{end_row}"
    
    def _freeze_panes(self, worksheet):
        """Freeze header row"""
        worksheet.freeze_panes = "A2"
    
    def _add_pie_chart(self, worksheet, category_data, chart_position="F5"):
        """Add pie chart for category breakdown with top categories only"""
        try:
            # Create simplified data for pie chart (top 8 categories + Others)
            chart_data = self._prepare_chart_data(category_data)
            
            # Write chart data to worksheet starting from column H
            start_row = 2
            for i, (category, amount) in enumerate(chart_data):
                worksheet.cell(row=start_row + i, column=8, value=category)  # Column H
                worksheet.cell(row=start_row + i, column=9, value=amount)    # Column I
            
            # Add headers for chart data
            worksheet.cell(row=1, column=8, value="Chart Categories")
            worksheet.cell(row=1, column=9, value="Amount (₹)")
            
            chart = PieChart()
            chart.title = "Top Spending Categories"
            chart.height = 12
            chart.width = 18
            
            # Data for chart using the simplified data
            data = Reference(worksheet, min_col=9, min_row=1, max_row=len(chart_data)+1, max_col=9)
            cats = Reference(worksheet, min_col=8, min_row=2, max_row=len(chart_data)+1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Add data labels with percentages
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = False  # Hide values, show only percentages
            
            worksheet.add_chart(chart, chart_position)
        except Exception as e:
            logger.warning(f"Could not add pie chart: {e}")
    
    def _prepare_chart_data(self, category_summary):
        """Prepare simplified data for pie chart - top 8 categories + Others"""
        # Sort by Total Debit descending
        sorted_data = category_summary.sort_values('Total Debit', ascending=False)
        
        # Take top 8 categories
        top_categories = sorted_data.head(8)
        remaining_categories = sorted_data.tail(len(sorted_data) - 8)
        
        chart_data = []
        
        # Add top categories
        for _, row in top_categories.iterrows():
            chart_data.append((row['Category'], row['Total Debit']))
        
        # Add "Others" if there are remaining categories
        if len(remaining_categories) > 0:
            others_total = remaining_categories['Total Debit'].sum()
            if others_total > 0:
                chart_data.append(("Others", others_total))
        
        return chart_data
    
    def _get_top_transactions(self, df, n=10):
        """Get top N transactions by amount"""
        # Combine debit and credit into single amount column
        df_copy = df.copy()
        df_copy['Amount'] = df_copy['Debit'].fillna(0) + df_copy['Credit'].fillna(0)
        return df_copy.nlargest(n, 'Amount')[['Txn Date', 'Description', 'Category', 'Amount']]
    
    def write_single_sheet(self, df: pd.DataFrame, sheet_name: str) -> None:
        """
        Write single DataFrame to Excel sheet with auto-fitted columns
        
        Args:
            df: DataFrame to write
            sheet_name: Name of the sheet
        """
        try:
            # Remove existing file if it exists (overwrite)
            if self.output_path.exists():
                self.output_path.unlink()
                
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Auto-fit columns
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set column width with padding, cap at 50 characters
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
            logger.info(f"Sheet '{sheet_name}' written to {self.output_path}")
        except Exception as e:
            logger.error(f"Error writing sheet '{sheet_name}': {e}")
            raise
    
    def _get_upi_construction_analysis(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate UPI construction spending analysis
        
        Args:
            df: Categorized transactions DataFrame
            
        Returns:
            DataFrame with UPI construction analysis
        """
        # Get all UPI categories
        upi_categories = df[df['Category'].str.startswith('UPI-', na=False)]
        
        if len(upi_categories) == 0:
            return pd.DataFrame()
        
        # Determine correct column names
        debit_col = 'Debit (₹)' if 'Debit (₹)' in df.columns else 'Debit'
        credit_col = 'Credit (₹)' if 'Credit (₹)' in df.columns else 'Credit'
        
        # Group by category with debit and credit analysis
        analysis = upi_categories.groupby('Category').agg({
            debit_col: ['sum', 'count'],
            credit_col: ['sum', 'count']
        }).round(2)
        
        # Flatten columns
        analysis.columns = ['Total Debit (₹)', 'Debit Count', 'Total Credit (₹)', 'Credit Count']
        analysis = analysis.reset_index()
        
        # Sort by highest debit value
        analysis = analysis.sort_values('Total Debit (₹)', ascending=False)
        
        # Add total row at the end (will be excluded from sorting)
        total_debit = analysis['Total Debit (₹)'].sum()
        total_debit_count = analysis['Debit Count'].sum()
        total_credit = analysis['Total Credit (₹)'].sum()
        total_credit_count = analysis['Credit Count'].sum()
        
        total_row = pd.DataFrame({
            'Category': ['TOTAL UPI'],
            'Total Debit (₹)': [total_debit],
            'Debit Count': [total_debit_count],
            'Total Credit (₹)': [total_credit],
            'Credit Count': [total_credit_count]
        })
        
        analysis = pd.concat([analysis, total_row], ignore_index=True)
        
        return analysis
    
    def _create_upi_hierarchical_sheet(self, df: pd.DataFrame, writer) -> None:
        """
        Create UPI summary sheet with hierarchical grouping and totals
        
        Args:
            df: Categorized transactions DataFrame
            writer: Excel writer object
        """
        # Filter UPI transactions
        upi_df = df[df['Category'].str.startswith('UPI-', na=False)].copy()
        
        if len(upi_df) == 0:
            return
        
        # Determine correct column name for amounts
        debit_col = 'Debit (₹)' if 'Debit (₹)' in upi_df.columns else 'Debit'
        
        # Create hierarchical summary
        summary_data = []
        
        # Group by category and calculate totals
        category_totals = upi_df.groupby('Category')[debit_col].sum().fillna(0)
        
        # Build hierarchy
        hierarchy = {}
        for category, total in category_totals.items():
            if not category.startswith('UPI-'):
                continue
                
            parts = category.split('-')
            level1 = parts[1] if len(parts) > 1 else 'Others'
            level2 = parts[2] if len(parts) > 2 else None
            
            if level1 not in hierarchy:
                hierarchy[level1] = {'total': 0, 'subcategories': {}}
            
            hierarchy[level1]['total'] += total
            
            if level2:
                if level2 not in hierarchy[level1]['subcategories']:
                    hierarchy[level1]['subcategories'][level2] = 0
                hierarchy[level1]['subcategories'][level2] += total
        
        # Build summary rows
        for level1, data in sorted(hierarchy.items()):
            # Add level 1 category
            summary_data.append({
                'Category': level1,
                'Total Amount (₹)': data['total'],
                'Level': 1
            })
            
            # Add level 2 subcategories
            for level2, total in sorted(data['subcategories'].items()):
                summary_data.append({
                    'Category': f"  {level2}",  # Indent for visual hierarchy
                    'Total Amount (₹)': total,
                    'Level': 2
                })
        
        # Create DataFrame
        summary_df = pd.DataFrame(summary_data)
        
        # Add total row - sum only Level 1 categories to avoid double counting
        level1_total = summary_df[summary_df['Level'] == 1]['Total Amount (₹)'].sum()
        total_row = pd.DataFrame({
            'Category': ['TOTAL UPI'],
            'Total Amount (₹)': [level1_total],
            'Level': [0]  # Special level for total
        })
        summary_df = pd.concat([summary_df, total_row], ignore_index=True)
        
        # Write to Excel
        summary_df[['Category', 'Total Amount (₹)']].to_excel(
            writer, sheet_name='UPI Summary', index=False
        )
        
        # Apply grouping
        workbook = writer.book
        worksheet = workbook['UPI Summary']
        
        self._apply_upi_summary_grouping(worksheet, summary_df)
        self._format_upi_summary(worksheet, summary_df)
    
    def _apply_upi_summary_grouping(self, worksheet, df: pd.DataFrame) -> None:
        """
        Apply Excel grouping to UPI summary with collapsed groups by default
        
        Args:
            worksheet: Excel worksheet object
            df: Summary DataFrame with Level column
        """
        level1_rows = []
        
        # Find all Level 1 category positions
        for idx, row in df.iterrows():
            if row['Level'] == 1:  # Level 1 category
                level1_rows.append(idx + 2)  # +2 for header and 1-based indexing
        
        # Group subcategories between Level 1 categories
        for i in range(len(level1_rows)):
            start_row = level1_rows[i]
            
            # Find end row (next Level 1 category or end of data, excluding total)
            if i + 1 < len(level1_rows):
                end_row = level1_rows[i + 1] - 1
            else:
                # Last category - find where Level 2 rows end (before total)
                end_row = start_row
                for idx in range(start_row - 1, len(df)):  # -1 to convert back to 0-based
                    if idx < len(df) and df.iloc[idx]['Level'] == 2:
                        end_row = idx + 2  # +2 for header and 1-based indexing
                    elif idx < len(df) and df.iloc[idx]['Level'] == 0:  # Total row
                        break
            
            # Group if there are subcategories
            if end_row > start_row:
                worksheet.row_dimensions.group(
                    start_row + 1, end_row, outline_level=1, hidden=True
                )
    
    def _format_upi_summary(self, worksheet, df: pd.DataFrame) -> None:
        """
        Apply visual formatting to UPI summary sheet
        
        Args:
            worksheet: Excel worksheet object
            df: Summary DataFrame with Level column
        """
        from openpyxl.styles import Font, PatternFill
        
        # Define styles
        level1_font = Font(bold=True, size=14)  # Bigger font
        level1_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        level2_font = Font(size=12)  # Bigger font
        total_font = Font(bold=True, size=16, color="FFFFFF")  # White text
        total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue background
        
        # Set row heights (bigger cells)
        for row_num in range(1, len(df) + 2):  # +2 for header and 1-based indexing
            worksheet.row_dimensions[row_num].height = 25
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 20
        
        for idx, row in df.iterrows():
            excel_row = idx + 2  # +2 for header and 1-based indexing
            level = row['Level']
            
            if level == 0:  # Total row
                worksheet.cell(row=excel_row, column=1).font = total_font
                worksheet.cell(row=excel_row, column=1).fill = total_fill
                worksheet.cell(row=excel_row, column=2).font = total_font
                worksheet.cell(row=excel_row, column=2).fill = total_fill
            elif level == 1:  # Level 1 (main categories) - Bold and colored
                worksheet.cell(row=excel_row, column=1).font = level1_font
                worksheet.cell(row=excel_row, column=1).fill = level1_fill
                worksheet.cell(row=excel_row, column=2).font = level1_font
                worksheet.cell(row=excel_row, column=2).fill = level1_fill
            else:  # Level 2 (subcategories) - Normal font
                worksheet.cell(row=excel_row, column=1).font = level2_font
                worksheet.cell(row=excel_row, column=2).font = level2_font
    
    def _format_upi_analysis_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        Format UPI Analysis sheet with TOTAL UPI row fixed at bottom using split panes
        
        Args:
            worksheet: Excel worksheet object
            df: UPI Analysis DataFrame
        """
        from openpyxl.styles import Font, PatternFill
        
        # Find TOTAL UPI row
        total_row_idx = None
        total_row_data = None
        for idx, row in df.iterrows():
            if row['Category'] == 'TOTAL UPI':
                total_row_idx = idx + 2  # +2 for header and 1-based indexing
                total_row_data = row
                break
        
        if total_row_idx and total_row_data is not None:
            # Calculate proper split position
            data_rows = len(df) - 1  # Exclude the TOTAL UPI row from count
            split_row = data_rows + 1  # Position for the split (after data, before total)
            
            # Remove TOTAL UPI from its current position
            worksheet.delete_rows(total_row_idx)
            
            # Add TOTAL UPI row at the bottom with some spacing
            total_position = split_row + 2  # Add some space before total
            
            # Create split panes - everything above split_row scrolls, total_position stays fixed
            worksheet.freeze_panes = f'A{split_row}'
            
            # Add TOTAL UPI row at the fixed bottom position
            total_font = Font(bold=True, size=14, color="FFFFFF")
            total_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            
            # Write TOTAL UPI data
            worksheet.cell(row=total_position, column=1, value="TOTAL UPI").font = total_font
            worksheet.cell(row=total_position, column=1).fill = total_fill
            
            # Add other columns from total_row_data
            col_names = df.columns.tolist()
            for col_idx, col_name in enumerate(col_names[1:], start=2):  # Skip 'Category' column
                cell = worksheet.cell(row=total_position, column=col_idx, value=total_row_data[col_name])
                cell.font = total_font
                cell.fill = total_fill
            
            # Set row height for better visibility
            worksheet.row_dimensions[total_position].height = 30
            
            # Add a thick border above the total row for separation
            from openpyxl.styles import Border, Side
            thick_border = Border(top=Side(style='thick'))
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=total_position, column=col).border = thick_border
    
    def _add_horizontal_bar_chart(self, worksheet, category_data, chart_position="J5"):
        """Add horizontal bar chart for better visualization when one category dominates"""
        try:
            from openpyxl.chart import BarChart, Reference
            
            # Create chart data excluding UPI to show other categories better
            chart_data = self._prepare_chart_data_for_bar(category_data)
            
            # Write chart data to worksheet starting from column J
            start_row = 2
            for i, (category, amount) in enumerate(chart_data):
                worksheet.cell(row=start_row + i, column=10, value=category)  # Column J
                worksheet.cell(row=start_row + i, column=11, value=amount)    # Column K
            
            # Add headers for chart data
            worksheet.cell(row=1, column=10, value="Chart Categories")
            worksheet.cell(row=1, column=11, value="Amount (₹)")
            
            # Create horizontal bar chart
            chart = BarChart()
            chart.type = "bar"  # Horizontal bars
            chart.title = "Non-UPI Spending Categories"
            chart.height = 12
            chart.width = 18
            chart.style = 10
            
            # Data for chart
            data = Reference(worksheet, min_col=11, min_row=1, max_row=len(chart_data)+1, max_col=11)
            categories = Reference(worksheet, min_col=10, min_row=2, max_row=len(chart_data)+1, max_col=10)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Add chart to worksheet
            worksheet.add_chart(chart, chart_position)
        except Exception as e:
            logger.warning(f"Could not add horizontal bar chart: {e}")
    
    def _prepare_chart_data_for_bar(self, category_summary):
        """Prepare data for bar chart - exclude UPI to show other categories better"""
        # Sort by Total Debit descending and exclude UPI
        debit_col = 'Total Debit (₹)' if 'Total Debit (₹)' in category_summary.columns else 'Total Debit'
        sorted_data = category_summary.sort_values(debit_col, ascending=False)
        non_upi_data = sorted_data[sorted_data['Category'] != 'UPI']
        
        # Take top 10 non-UPI categories
        top_categories = non_upi_data.head(10)
        
        chart_data = []
        for _, row in top_categories.iterrows():
            chart_data.append((row['Category'], row[debit_col]))
        
        return chart_data
    
    def _create_filter_controls_sheet(self, df: pd.DataFrame, writer) -> None:
        """
        Create Filter Controls sheet with Year/Month dropdowns for time range filtering
        
        Args:
            df: Categorized transactions DataFrame
            writer: Excel writer object
        """
        # Extract available years and months from data
        try:
            # Handle different date formats
            if df['Txn Date'].dtype == 'object':
                # Try to parse as datetime with mixed formats
                df['ParsedDate'] = pd.to_datetime(df['Txn Date'], format='mixed', dayfirst=True)
            else:
                df['ParsedDate'] = pd.to_datetime(df['Txn Date'])
            
            df['Year'] = df['ParsedDate'].dt.year
            df['Month'] = df['ParsedDate'].dt.month
        except:
            # Fallback: extract year from string
            df['Year'] = df['Txn Date'].astype(str).str.extract(r'(\d{4})')[0].astype(int, errors='ignore')
            df['Month'] = df['Txn Date'].astype(str).str.extract(r'\d{2}-(\d{2})-\d{4}')[0].astype(int, errors='ignore')
        
        available_years = sorted(df['Year'].dropna().unique())
        available_months = list(range(1, 13))  # 1-12 for all months
        
        # Create filter controls data
        controls_data = []
        
        # Header
        controls_data.append(['FINANCE ANALYZER - TIME RANGE FILTER', '', '', ''])
        controls_data.append(['', '', '', ''])
        
        # Filter controls
        controls_data.append(['Filter by Year-Month:', '', '', ''])
        controls_data.append(['', '', '', ''])
        controls_data.append(['Selected Year:', available_years[0] if available_years else 2023, '', ''])
        controls_data.append(['Selected Month:', 1, '', ''])
        controls_data.append(['', '', '', ''])
        
        # Available options (for reference)
        controls_data.append(['Available Years:', ', '.join(map(str, available_years)), '', ''])
        controls_data.append(['Available Months:', '01-Jan, 02-Feb, 03-Mar, 04-Apr, 05-May, 06-Jun', '', ''])
        controls_data.append(['', '07-Jul, 08-Aug, 09-Sep, 10-Oct, 11-Nov, 12-Dec', '', ''])
        controls_data.append(['', '', '', ''])
        
        # Instructions
        controls_data.append(['INSTRUCTIONS:', '', '', ''])
        controls_data.append(['1. Change Selected Year (B5) to filter by year', '', '', ''])
        controls_data.append(['2. Change Selected Month (B6) to filter by month', '', '', ''])
        controls_data.append(['3. Use 0 for "All Months" in a year', '', '', ''])
        controls_data.append(['4. All other sheets will update automatically', '', '', ''])
        controls_data.append(['', '', '', ''])
        
        # Current status
        total_transactions = len(df)
        date_range = f"{df['Txn Date'].min()} to {df['Txn Date'].max()}"
        controls_data.append(['CURRENT STATUS:', '', '', ''])
        controls_data.append(['Total Transactions:', total_transactions, '', ''])
        controls_data.append(['Date Range:', date_range, '', ''])
        
        # Create DataFrame and write to Excel
        controls_df = pd.DataFrame(controls_data, columns=['Parameter', 'Value', 'Col3', 'Col4'])
        controls_df.to_excel(writer, sheet_name='Filter Controls', index=False, header=False)
        
        # Format the Filter Controls sheet
        workbook = writer.book
        worksheet = workbook['Filter Controls']
        self._format_filter_controls(worksheet, available_years, available_months)
    
    def _format_filter_controls(self, worksheet, available_years, available_months) -> None:
        """
        Apply formatting and data validation to Filter Controls sheet
        
        Args:
            worksheet: Excel worksheet object
            available_years: List of available years
            available_months: List of available months (1-12)
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Define styles
        header_font = Font(bold=True, size=16, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        label_font = Font(bold=True, size=12)
        value_font = Font(size=11)
        instruction_font = Font(size=10, italic=True)
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        
        # Format header
        worksheet.merge_cells('A1:D1')
        worksheet.cell(row=1, column=1).font = header_font
        worksheet.cell(row=1, column=1).fill = header_fill
        worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # Format labels and values
        for row in range(3, 12):  # Filter controls section
            worksheet.cell(row=row, column=1).font = label_font
            worksheet.cell(row=row, column=2).font = value_font
        
        # Format instructions
        for row in range(13, 18):  # Instructions section
            worksheet.cell(row=row, column=1).font = instruction_font
            worksheet.cell(row=row, column=2).font = instruction_font
        
        # Add data validation for Year dropdown (B5)
        year_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(map(str, available_years))}"',
            allow_blank=False
        )
        year_validation.add('B5')
        worksheet.add_data_validation(year_validation)
        
        # Add data validation for Month dropdown (B6)
        month_options = ["0-All"] + [f"{i:02d}-{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][i-1]}" for i in range(1, 13)]
        month_validation = DataValidation(
            type="list", 
            formula1=f'"{",".join(month_options)}"',
            allow_blank=False
        )
        month_validation.add('B6')
        worksheet.add_data_validation(month_validation)
        
        # Set default values
        worksheet.cell(row=6, column=2).value = "0-All"  # Default to "All Months"

    def append_sheet_to_existing(self, df: pd.DataFrame, sheet_name: str) -> None:
        """
        Append a new sheet to existing Excel file
        
        Args:
            df: DataFrame to write
            sheet_name: Name of the new sheet
        """
        try:
            if self.output_path.exists():
                with pd.ExcelWriter(
                    self.output_path, 
                    engine='openpyxl', 
                    mode='a', 
                    if_sheet_exists='replace'
                ) as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Create new file if it doesn't exist
                self.write_single_sheet(df, sheet_name)
                
            logger.info(f"Sheet '{sheet_name}' appended to {self.output_path}")
            
        except Exception as e:
            logger.error(f"Error appending sheet '{sheet_name}': {e}")
            raise
